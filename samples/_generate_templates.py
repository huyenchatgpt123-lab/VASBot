"""One-shot script to generate Excel import templates. Run: python _generate_templates.py"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

out = Path(__file__).resolve().parent
thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
header_fill = PatternFill("solid", fgColor="1E40AF")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
cell_font = Font(name="Calibri", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, cols):
    for col, title in enumerate(cols, 1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin


def style_rows(ws, n_rows, n_cols, left_cols=(1, 2)):
    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            cell.font = cell_font
            cell.border = thin
            cell.alignment = left if c in left_cols else center


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_notes(ws, lines):
    for i, line in enumerate(lines, 1):
        cell = ws.cell(i, 1, line)
        bold = i == 1 or line.startswith(("QUY", "CỘT", "KIỂM", "HƯỚNG", "1.", "2.", "3."))
        cell.font = Font(bold=bold or (i == 1), name="Calibri", size=12 if i == 1 else 11)
    ws.column_dimensions["A"].width = 100


# ---------- File 1: Danh sách giáo viên ----------
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Giao vien"

headers1 = [
    "Họ tên", "Email", "Mật khẩu", "Vai trò", "Phòng ban",
    "Biệt danh", "Chức vụ", "Mã GV", "Cơ sở",
]
style_header(ws1, headers1)

rows1 = [
    ["Trần Chí Hải", "hai.tran@vietanhschool.edu.vn", "Abc@12345", "user", "Tổ Tin học", "Hải Tin", "Giáo viên", "GV001", "VA1"],
    ["Nguyễn Thị Thế Đoan", "doan.nguyen@vietanhschool.edu.vn", "Abc@12345", "user", "Tổ Toán", "Thế Đoan", "Giáo viên", "GV002", "VA1"],
    ["Lê Văn Minh", "minh.le@vietanhschool.edu.vn", "Abc@12345", "user", "Tổ Ngữ Văn", "Minh Văn", "Tổ trưởng", "GV003", "VA3"],
    ["Phạm Thu Hà", "ha.pham@vietanhschool.edu.vn", "Abc@12345", "user", "Tổ Tiếng Anh", "Hà Anh", "Giáo viên", "GV004", "EMC"],
]
for i, row in enumerate(rows1, 2):
    for c, val in enumerate(row, 1):
        ws1.cell(i, c, val)
style_rows(ws1, 1 + len(rows1), len(headers1), left_cols=(1, 2, 5, 6, 7))
autosize(ws1, [22, 34, 14, 10, 16, 14, 12, 10, 10])
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:I{1 + len(rows1)}"

dv_role = DataValidation(type="list", formula1='"admin,user"', allow_blank=True)
dv_campus = DataValidation(type="list", formula1='"VA1,VA3,EMC"', allow_blank=True)
ws1.add_data_validation(dv_role)
ws1.add_data_validation(dv_campus)
dv_role.add("D2:D500")
dv_campus.add("I2:I500")

ws1n = wb1.create_sheet("Huong dan")
write_notes(ws1n, [
    "HƯỚNG DẪN ĐIỀN DANH SÁCH GIÁO VIÊN",
    "",
    "1. Giữ nguyên dòng tiêu đề (dòng 1). Không đổi tên cột.",
    "2. Cột bắt buộc: Họ tên, Email, Mật khẩu.",
    "3. Mã GV (cột H): dạng GV001, GV002... — không dấu, không khoảng trắng, không đổi sau khi gán.",
    "4. Phòng ban = tổ chuyên môn (Tổ Toán, Tổ Tin học, Tổ Ngữ Văn...).",
    "5. Biệt danh: nên điền nếu thời khóa biểu ghi tên rút gọn.",
    "6. Cơ sở: VA1 / VA3 / EMC — có thể để trống, hệ thống suy ra từ TKB sau.",
    "7. Vai trò: admin hoặc user (để trống = user).",
    "8. Không gộp ô. Không để dòng trống xen giữa dữ liệu.",
    "9. Không cần cột môn — hệ thống chỉ quan tâm tổ và lịch trống.",
])

path1 = out / "mau_danh_sach_giao_vien.xlsx"
wb1.save(path1)

# ---------- File 2: Thời khóa biểu bảng phẳng ----------
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Thoi khoa bieu"

headers2 = ["Mã GV", "Họ tên", "Cơ sở", "Thứ", "Tiết", "Lớp"]
style_header(ws2, headers2)

hai = "Trần Chí Hải"
code = "GV001"
campus = "VA1"
# period 1-8 global. Sample matches PDF morning slots (periods 1-5).
slots = [
    (2, 2, "6A6"),
    (2, 3, "6A6"),
    (2, 4, "8C2"),
    (2, 5, "8C2"),
    (3, 4, "6A8"),
    (3, 5, "6A8"),
    (4, 1, "6A7"),
    (4, 2, "6A7"),
    (5, 2, "8C7"),
    (5, 3, "8C7"),
    (6, 1, "8C3"),
    (6, 2, "8C3"),
    (6, 3, "6A5"),
    (6, 4, "6A5"),
]
doan = [
    (2, 1, "7A1"),
    (2, 2, "7A1"),
    (3, 6, "8A2"),  # chiều tiết 1 = period 6
    (5, 3, "9A1"),
]

r = 2
for thu, tiet, lop in slots:
    ws2.cell(r, 1, code)
    ws2.cell(r, 2, hai)
    ws2.cell(r, 3, campus)
    ws2.cell(r, 4, thu)
    ws2.cell(r, 5, tiet)
    ws2.cell(r, 6, lop)
    r += 1

for thu, tiet, lop in doan:
    ws2.cell(r, 1, "GV002")
    ws2.cell(r, 2, "Nguyễn Thị Thế Đoan")
    ws2.cell(r, 3, "VA1")
    ws2.cell(r, 4, thu)
    ws2.cell(r, 5, tiet)
    ws2.cell(r, 6, lop)
    r += 1

last = r - 1
style_rows(ws2, last, len(headers2), left_cols=(2,))
autosize(ws2, [10, 22, 10, 8, 8, 10])
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:F{last}"

dv_campus2 = DataValidation(type="list", formula1='"VA1,VA3,EMC"', allow_blank=False)
dv_thu = DataValidation(type="list", formula1='"2,3,4,5,6,7"', allow_blank=False)
ws2.add_data_validation(dv_campus2)
ws2.add_data_validation(dv_thu)
dv_campus2.add("C2:C2000")
dv_thu.add("D2:D2000")

ws2n = wb2.create_sheet("Huong dan")
write_notes(ws2n, [
    "HƯỚNG DẪN ĐIỀN THỜI KHÓA BIỂU (BẢNG PHẲNG)",
    "",
    "QUY TẮC VÀNG",
    "- Mỗi tiết = 1 dòng. Tiết đôi viết thành 2 dòng.",
    "- Không gộp ô. Không chèn dòng tiêu đề phụ giữa bảng.",
    "",
    "CỘT",
    "- Mã GV: khớp với file danh sách giáo viên.",
    "- Họ tên: không viết Thầy/Cô; dùng khi thiếu mã.",
    "- Cơ sở: VA1 / VA3 / EMC.",
    "- Thứ: số 2 đến 7.",
    "- Tiết: 1–8 cả ngày (1–5 sáng, 6–8 chiều).",
    "- Lớp: chỉ tên lớp (6A7).",
    "",
    "Import mới sẽ THAY toàn bộ TKB hiện hành của cơ sở đã chọn.",
    "",
    "KIỂM TRA",
    "1. Đếm dòng theo Mã GV so với Số tiết trên PDF gốc.",
    "2. Không trùng (Mã GV + Thứ + Tiết).",
])

path2 = out / "mau_thoi_khoa_bieu.xlsx"
wb2.save(path2)

print(path1)
print(path2)
print(f"teachers={len(rows1)} timetable_rows={last - 1}")
