"""Parse timetable Excel — flat table or school grid layout.

Flat columns:
  Mã GV | Họ tên | Cơ sở | Thứ | Tiết | Lớp

Grid layout (2 header rows + data):
  A Mã GV | B Cơ sở | C Giáo viên | D Buổi dạy | Thứ 2..6 × tiết 1–5
  Sáng tiết 1–5 → period 1–5; Chiều tiết 1–3 → period 6–8; Chiều 4–5 bỏ qua.
"""
from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


FIELD_ALIASES: Dict[str, List[str]] = {
    "teacher_code": ["ma gv", "magv", "ma giao vien", "teacher code", "teacher_code", "ma"],
    "name": ["ho ten", "hoten", "ten", "name", "ho va ten", "giao vien"],
    "campus": ["co so", "coso", "campus", "co so"],
    "day": ["thu", "day", "day_of_week", "thu trong tuan"],
    "period": ["tiet", "period", "tiet hoc"],
    "class_name": ["lop", "class", "lop hoc", "ten lop"],
    "session": ["buoi day", "buoi", "session", "buoi hoc"],
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(c for c in text if unicodedata.category(c) != "Mn")


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _guess_grade(class_name: str) -> Optional[int]:
    m = re.match(r"^(\d{1,2})", class_name.strip())
    if not m:
        return None
    grade = int(m.group(1))
    return grade if 1 <= grade <= 12 else None


def _strip_honorific(name: str) -> str:
    for prefix in ("Thầy ", "Cô ", "thầy ", "cô "):
        if name.startswith(prefix):
            return name[len(prefix):].strip()
    return name


def build_column_map(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    normalized_aliases = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in FIELD_ALIASES.items()
    }
    column_map: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        header = _normalize_header(cell)
        if not header:
            continue
        for field, aliases in normalized_aliases.items():
            if header in aliases and field not in column_map:
                column_map[field] = idx
                break
    return column_map


def _is_flat_format(header_row: Tuple[Any, ...]) -> bool:
    column_map = build_column_map(header_row)
    return (
        "day" in column_map
        and "period" in column_map
        and "class_name" in column_map
    )


def _parse_thu_number(text: str) -> Optional[int]:
    h = _normalize_header(text)
    if not h:
        return None
    m = re.search(r"thu\s*(\d)", h)
    if m:
        n = int(m.group(1))
        return n if 2 <= n <= 6 else None
    if h.isdigit():
        n = int(h)
        return n if 2 <= n <= 6 else None
    return None


def _is_period_row(row: Tuple[Any, ...]) -> bool:
    nums = 0
    for cell in row[4:] if len(row) > 4 else row:
        if _cell_to_str(cell) in ("1", "2", "3", "4", "5"):
            nums += 1
    return nums >= 8


def _find_grid_layout(all_rows: List[Tuple[Any, ...]]) -> Optional[Tuple[int, int, int, Dict[str, int]]]:
    """Return (header_idx, period_idx, data_start, fixed_col_map) or None."""
    for i in range(min(6, len(all_rows))):
        row = all_rows[i]
        col_map = build_column_map(row)
        if "teacher_code" not in col_map:
            continue
        period_idx = i + 1 if i + 1 < len(all_rows) and _is_period_row(all_rows[i + 1]) else None
        if period_idx is None:
            continue
        if "session" in col_map or any(
            _normalize_header(c) in ("buoi day", "buoi")
            for c in row
            if c is not None
        ):
            return i, period_idx, period_idx + 1, col_map
    return None


def _build_grid_schedule_columns(
    header_row: Tuple[Any, ...],
    period_row: Tuple[Any, ...],
    fixed_end_col: int,
) -> List[Tuple[int, int, int]]:
    """(col_index, day_of_week, period_in_day)."""
    max_len = max(len(header_row), len(period_row))
    cols: List[Tuple[int, int, int]] = []
    current_day: Optional[int] = None

    for col in range(fixed_end_col, max_len):
        if col < len(header_row):
            day_from_header = _parse_thu_number(_cell_to_str(header_row[col]))
            if day_from_header:
                current_day = day_from_header

        if current_day is None:
            continue

        if col >= len(period_row):
            continue
        p_raw = _cell_to_str(period_row[col])
        if not p_raw:
            continue
        try:
            p_in_day = int(float(p_raw))
        except ValueError:
            continue
        if 1 <= p_in_day <= 5:
            cols.append((col, current_day, p_in_day))
    return cols


def _global_period_from_session(session: str, period_in_day: int) -> Optional[int]:
    s = _normalize_header(session)
    if "chieu" in s:
        if period_in_day >= 4:
            return None
        return period_in_day + 5
    if 1 <= period_in_day <= 5:
        return period_in_day
    return None


def _parse_flat_rows(all_rows: List[Tuple[Any, ...]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    errors: List[str] = []
    column_map = build_column_map(all_rows[0])
    if "teacher_code" not in column_map and "name" not in column_map:
        return [], ["Cần ít nhất một trong hai cột: Mã GV hoặc Họ tên"]

    rows: List[Dict[str, Any]] = []
    seen_teacher_slot: set = set()

    for i, row in enumerate(all_rows[1:], start=2):
        if not row or all(cell is None or _cell_to_str(cell) == "" for cell in row):
            continue

        def get(field: str) -> str:
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _cell_to_str(row[idx])

        teacher_code = get("teacher_code").upper()
        name = _strip_honorific(get("name"))
        campus = get("campus").upper()
        day_raw = get("day")
        period_raw = get("period")
        class_name = get("class_name")

        try:
            day = int(float(day_raw)) if day_raw else 0
        except ValueError:
            errors.append(f"Dòng {i}: Thứ không hợp lệ '{day_raw}'")
            continue
        try:
            period = int(float(period_raw)) if period_raw else 0
        except ValueError:
            errors.append(f"Dòng {i}: Tiết không hợp lệ '{period_raw}'")
            continue

        if day < 2 or day > 7:
            errors.append(f"Dòng {i}: Thứ phải từ 2 đến 7 (got {day})")
            continue
        if period < 1 or period > 8:
            errors.append(f"Dòng {i}: Tiết phải từ 1 đến 8 (got {period})")
            continue
        if not class_name:
            errors.append(f"Dòng {i}: thiếu Lớp")
            continue
        if not teacher_code and not name:
            errors.append(f"Dòng {i}: thiếu Mã GV và Họ tên")
            continue

        key = (teacher_code or name.lower(), day, period)
        if key in seen_teacher_slot:
            errors.append(f"Dòng {i}: trùng tiết của cùng giáo viên (Thứ {day} tiết {period})")
            continue
        seen_teacher_slot.add(key)

        rows.append({
            "row": i,
            "teacher_code": teacher_code or None,
            "name": name or None,
            "campus": campus or None,
            "day": day,
            "period": period,
            "class_name": class_name,
            "grade": _guess_grade(class_name),
        })

    return rows, errors


def _parse_grid_rows(all_rows: List[Tuple[Any, ...]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    layout = _find_grid_layout(all_rows)
    if not layout:
        return [], ["Không nhận dạng được lưới thời khóa biểu (cần Mã GV, Buổi dạy, hàng tiết 1–5)"]

    header_idx, period_idx, data_start, col_map = layout
    header_row = all_rows[header_idx]
    period_row = all_rows[period_idx]

    code_col = col_map.get("teacher_code", 0)
    campus_col = col_map.get("campus", 1)
    name_col = col_map.get("name", 2)
    session_col = col_map.get("session", 3)
    fixed_end = max(code_col, campus_col, name_col, session_col) + 1

    schedule_cols = _build_grid_schedule_columns(header_row, period_row, fixed_end)
    if not schedule_cols:
        return [], ["Không tìm thấy cột Thứ / Tiết trong lưới TKB"]

    errors: List[str] = []
    rows: List[Dict[str, Any]] = []
    seen_teacher_slot: set = set()

    last_code = ""
    last_name = ""
    last_campus = ""

    for i, row in enumerate(all_rows[data_start:], start=data_start + 1):
        if not row or all(cell is None or _cell_to_str(cell) == "" for cell in row):
            continue

        def at(col: int) -> str:
            if col >= len(row):
                return ""
            return _cell_to_str(row[col])

        teacher_code = at(code_col).upper()
        name = _strip_honorific(at(name_col))
        campus = at(campus_col).upper()
        session = at(session_col)

        if teacher_code:
            last_code = teacher_code
        else:
            teacher_code = last_code
        if name:
            last_name = name
        else:
            name = last_name
        if campus:
            last_campus = campus
        else:
            campus = last_campus

        session_norm = _normalize_header(session)
        if not session_norm or ("sang" not in session_norm and "chieu" not in session_norm):
            continue
        if not teacher_code and not name:
            continue

        for col_idx, day, period_in_day in schedule_cols:
            class_name = at(col_idx)
            if not class_name:
                continue

            period = _global_period_from_session(session, period_in_day)
            if period is None:
                continue

            key = (teacher_code or name.lower(), day, period)
            if key in seen_teacher_slot:
                errors.append(
                    f"Dòng {i}: trùng tiết của cùng giáo viên (Thứ {day} tiết {period})"
                )
                continue
            seen_teacher_slot.add(key)

            rows.append({
                "row": i,
                "teacher_code": teacher_code or None,
                "name": name or None,
                "campus": campus or None,
                "day": day,
                "period": period,
                "class_name": class_name,
                "grade": _guess_grade(class_name),
            })

    if not rows and not errors:
        return [], ["Không có tiết nào trong lưới TKB — kiểm tra ô lớp và buổi Sáng/Chiều"]
    return rows, errors


def parse_timetable_excel(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Return (rows, errors). Each row: teacher_code, name, campus, day, period, class_name, grade."""
    try:
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return [], ["File Excel không hợp lệ"]

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], ["File Excel trống"]

    if _is_flat_format(all_rows[0]):
        return _parse_flat_rows(all_rows)

    grid_rows, grid_errors = _parse_grid_rows(all_rows)
    if grid_rows:
        return grid_rows, grid_errors

    if _find_grid_layout(all_rows):
        return grid_rows, grid_errors

    return [], [
        "Không nhận dạng được file TKB. "
        "Dùng bảng phẳng (Mã GV, Cơ sở, Thứ, Tiết, Lớp) hoặc lưới (Mã GV, Cơ sở, Giáo viên, Buổi dạy, Thứ 2–6)."
    ]
