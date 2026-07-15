import io
import logging

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from sqlalchemy.orm import Session
from typing import List, Optional
from openpyxl import load_workbook

from app.database import get_db
from app.schemas.auth import UserResponse, UserCreate, UserUpdate
from app.schemas.dashboard import DashboardResponse
from app.services.dashboard_service import DashboardService
from app.services.auth_service import AuthService
from app.services.task_service import TaskService
from app.repositories.user_repository import UserRepository
from app.repositories.position_repository import PositionRepository
from app.repositories.department_repository import DepartmentRepository
from app.utils.auth import require_admin, hash_password
from app.utils.excel_user_import import build_column_map, parse_user_row, is_empty_row
from app.utils.user_serializer import serialize_user
from app.models.user import User
from app.schemas.position import PositionCreate, PositionUpdate, PositionResponse
from app.schemas.department import DepartmentCreate, DepartmentUpdate, DepartmentResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin", tags=["Admin"])


@router.get("/dashboard", response_model=DashboardResponse)
def get_dashboard(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    service = DashboardService(db)
    data = service.get_dashboard(start_date=start_date, end_date=end_date)
    return DashboardResponse(**data)


@router.get("/users", response_model=List[UserResponse])
def get_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    repo = UserRepository(db)
    users = repo.get_all()
    return [serialize_user(u) for u in users]


@router.post("/users", response_model=UserResponse)
def create_user(
    data: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    service = AuthService(db)
    if not data.department_id and not data.department:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Phòng ban là bắt buộc")
    try:
        user = service.create_user(data, require_nickname=False)
        TaskService(db).rematch_assignees(user_id=user.id)
        return serialize_user(user)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/users/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    repo = UserRepository(db)
    user = repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Người dùng không tồn tại")

    update_fields = {}
    if data.name is not None:
        update_fields["name"] = data.name
    if data.email is not None:
        existing = repo.get_by_email(data.email)
        if existing and existing.id != user_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email đã được sử dụng")
        update_fields["email"] = data.email
    if data.password is not None:
        update_fields["password_hash"] = hash_password(data.password)
    if data.role is not None:
        update_fields["role"] = data.role
    if data.department_id is not None:
        dept_repo = DepartmentRepository(db)
        dept = dept_repo.get_by_id(data.department_id)
        if not dept:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Phòng ban không tồn tại")
        update_fields["department_id"] = dept.id
        update_fields["department"] = dept.name
    elif data.department is not None:
        dept_repo = DepartmentRepository(db)
        dept = dept_repo.resolve_by_name(data.department) if data.department else None
        update_fields["department_id"] = dept.id if dept else None
        update_fields["department"] = dept.name if dept else None
    if data.position_id is not None:
        pos_repo = PositionRepository(db)
        pos = pos_repo.get_by_id(data.position_id)
        if not pos:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chức vụ không tồn tại")
        update_fields["position_id"] = pos.id
        update_fields["position"] = pos.name
    elif data.position is not None:
        pos_repo = PositionRepository(db)
        pos = pos_repo.resolve_by_name(data.position) if data.position else None
        update_fields["position_id"] = pos.id if pos else None
        update_fields["position"] = pos.name if pos else None
    if data.nickname is not None:
        nickname = data.nickname.strip()
        if not nickname:
            update_fields["nickname"] = None
        else:
            if repo.nickname_exists(nickname, exclude_id=user_id):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Biệt danh '{nickname}' đã được sử dụng",
                )
            update_fields["nickname"] = nickname

    updated = repo.update(user_id, **update_fields)
    if "nickname" in update_fields:
        TaskService(db).rematch_assignees(user_id=user_id)
    return serialize_user(updated)


@router.get("/positions", response_model=List[PositionResponse])
def get_positions(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    pos_repo = PositionRepository(db)
    positions = pos_repo.get_all()
    return [
        PositionResponse(
            id=p.id,
            name=p.name,
            can_upload=p.can_upload,
            can_manage_tasks=p.can_manage_tasks,
            can_delete_documents=p.can_delete_documents,
            scope_all_departments=p.scope_all_departments,
            sort_order=p.sort_order,
            user_count=pos_repo.count_users(p.id),
        )
        for p in positions
    ]


@router.post("/positions", response_model=PositionResponse)
def create_position(
    data: PositionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    pos_repo = PositionRepository(db)
    if pos_repo.get_by_name(data.name):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chức vụ đã tồn tại")
    pos = pos_repo.create(**data.model_dump())
    return PositionResponse(
        id=pos.id,
        name=pos.name,
        can_upload=pos.can_upload,
        can_manage_tasks=pos.can_manage_tasks,
        can_delete_documents=pos.can_delete_documents,
        scope_all_departments=pos.scope_all_departments,
        sort_order=pos.sort_order,
        user_count=0,
    )


@router.put("/positions/{position_id}", response_model=PositionResponse)
def update_position(
    position_id: int,
    data: PositionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    pos_repo = PositionRepository(db)
    if data.name:
        existing = pos_repo.get_by_name(data.name)
        if existing and existing.id != position_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chức vụ đã tồn tại")
    pos = pos_repo.update(position_id, **data.model_dump(exclude_unset=True))
    if not pos:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Chức vụ không tồn tại")
    users_with_pos = db.query(User).filter(User.position_id == position_id).all()
    for u in users_with_pos:
        u.position = pos.name
    db.commit()
    return PositionResponse(
        id=pos.id,
        name=pos.name,
        can_upload=pos.can_upload,
        can_manage_tasks=pos.can_manage_tasks,
        can_delete_documents=pos.can_delete_documents,
        scope_all_departments=pos.scope_all_departments,
        sort_order=pos.sort_order,
        user_count=pos_repo.count_users(pos.id),
    )


@router.delete("/positions/{position_id}")
def delete_position(
    position_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    pos_repo = PositionRepository(db)
    try:
        if not pos_repo.delete(position_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Chức vụ không tồn tại")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return {"message": "Chức vụ đã được xóa"}


@router.get("/departments", response_model=List[DepartmentResponse])
def get_departments(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    dept_repo = DepartmentRepository(db)
    departments = dept_repo.get_all()
    return [
        DepartmentResponse(
            id=d.id,
            name=d.name,
            sort_order=d.sort_order,
            user_count=dept_repo.count_users(d.id),
        )
        for d in departments
    ]


@router.post("/departments", response_model=DepartmentResponse)
def create_department(
    data: DepartmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    dept_repo = DepartmentRepository(db)
    if dept_repo.get_by_name(data.name):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Phòng ban đã tồn tại")
    dept = dept_repo.create(**data.model_dump())
    return DepartmentResponse(
        id=dept.id, name=dept.name, sort_order=dept.sort_order, user_count=0,
    )


@router.put("/departments/{department_id}", response_model=DepartmentResponse)
def update_department(
    department_id: int,
    data: DepartmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    dept_repo = DepartmentRepository(db)
    if data.name:
        existing = dept_repo.get_by_name(data.name)
        if existing and existing.id != department_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Phòng ban đã tồn tại")
    dept = dept_repo.update(department_id, **data.model_dump(exclude_unset=True))
    if not dept:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Phòng ban không tồn tại")
    return DepartmentResponse(
        id=dept.id,
        name=dept.name,
        sort_order=dept.sort_order,
        user_count=dept_repo.count_users(dept.id),
    )


@router.delete("/departments/{department_id}")
def delete_department(
    department_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    dept_repo = DepartmentRepository(db)
    try:
        if not dept_repo.delete(department_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Phòng ban không tồn tại")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return {"message": "Phòng ban đã được xóa"}


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    if user_id == current_user.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không thể xóa tài khoản của bạn")

    repo = UserRepository(db)
    try:
        if not repo.delete(user_id, reassign_documents_to=current_user.id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Người dùng không tồn tại")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    return {"message": "Người dùng đã được xóa"}


@router.post("/users/import-excel")
async def import_users_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chỉ chấp nhận file Excel (.xlsx)")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File Excel không hợp lệ")

    service = AuthService(db)
    repo = UserRepository(db)
    created = 0
    skipped = 0
    errors = []

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File Excel trống")

    column_map = build_column_map(all_rows[0])
    data_rows = all_rows[1:]
    seen_nicknames: set[str] = set()

    for i, row in enumerate(data_rows, start=2):
        if is_empty_row(row):
            continue

        parsed = parse_user_row(row, column_map)
        name = parsed["name"] or ""
        email = parsed["email"] or ""
        password = parsed["password"] or ""
        nickname = (parsed["nickname"] or "").strip()

        if not name or not email or not password:
            errors.append(f"Dòng {i}: thiếu họ tên, email hoặc mật khẩu")
            skipped += 1
            continue

        if nickname:
            nickname_key = nickname.lower()
            if nickname_key in seen_nicknames:
                errors.append(f"Dòng {i}: biệt danh '{nickname}' bị trùng trong file")
                skipped += 1
                continue
            seen_nicknames.add(nickname_key)

            if repo.nickname_exists(nickname):
                errors.append(f"Dòng {i}: biệt danh '{nickname}' đã tồn tại")
                skipped += 1
                continue

        existing = repo.get_by_email(email)
        if existing:
            errors.append(f"Dòng {i}: email {email} đã tồn tại")
            skipped += 1
            continue

        try:
            user_data = UserCreate(
                name=name,
                nickname=nickname or None,
                email=email,
                password=password,
                role=parsed["role"] or "user",
                department=parsed["department"],
                position=parsed["position"],
            )
            new_user = service.create_user(user_data, require_nickname=False)
            TaskService(db).rematch_assignees(user_id=new_user.id)
            created += 1
        except Exception as e:
            errors.append(f"Dòng {i}: {str(e)}")
            skipped += 1

    return {
        "message": f"Import hoàn tất: {created} người dùng được tạo, {skipped} bị bỏ qua",
        "created": created,
        "skipped": skipped,
        "errors": errors[:50],
    }
